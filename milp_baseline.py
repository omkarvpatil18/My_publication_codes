"""
MILP Baseline for ASP Comparison
Paper: A Logic-driven Integrated System for Resilient and Sustainable
       Supplier Selection in Supply Chains
Solver: scipy.optimize.milp (HiGHS backend, open-source)
Dataset: Figshare DOI 10.6084/m9.figshare.31305760

All three scales now use Excel files as the primary data source
(dataset_100.xlsx, dataset_500.xlsx, dataset_1000.xlsx) with
disruption assignments from the companion .lp files.

Implements Equations (1)-(8) from Section 4.1.
"""

import re
import time
import json
import numpy as np
import openpyxl
from scipy.optimize import milp, LinearConstraint, Bounds
from scipy.sparse import csr_matrix


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — DATASET PARSER (Excel + .lp)
# Used for all three scales: 100, 500, 1000 suppliers
# ─────────────────────────────────────────────────────────────────────────────

def parse_excel_lp(xlsx_path, lp_path):
    """
    Parse supplier data from Excel workbook and disruptions from .lp file.

    Excel sheets used:
        'Primary Suppliers' : supplier ID, capacity, cost, social score, env score
        'Backup Suppliers'  : same structure for backup suppliers
        'Policy Parameters' : demand, ESG thresholds, weights

    .lp file used for:
        disrupted(S, W) facts — which supplier is disrupted in which scenario
        scenario(W)     facts — scenario names
        soc_threshold / env_threshold — override Excel if present
    """
    wb     = openpyxl.load_workbook(xlsx_path, data_only=True)
    lp_txt = open(lp_path).read()

    def read_sheet(sheet, cols, skip=2):
        """Read a supplier sheet, skipping header rows and AVERAGE summary row."""
        ws   = wb[sheet]
        rows = []
        for row in ws.iter_rows(min_row=skip + 1, values_only=True):
            # Skip empty rows, header row, and the AVERAGE summary row
            if row[0] and row[0] not in ('Supplier ID', 'AVERAGE'):
                rows.append([row[c] for c in cols])
        return rows

    # cols order: [ID, capacity, unit_cost, social_score, env_score]
    prim = read_sheet('Primary Suppliers', [0, 1, 2, 3, 4])
    back = read_sheet('Backup Suppliers',  [0, 1, 2, 3, 4])

    primaries = [r[0] for r in prim]
    backups   = [r[0] for r in back]
    scenarios = sorted(set(re.findall(r'scenario\((\w+)\)', lp_txt)))

    # Index maps: supplier_id -> array index
    pi = {s: i for i, s in enumerate(primaries)}
    bi = {b: i for i, b in enumerate(backups)}
    si = {w: i for i, w in enumerate(scenarios)}
    nP, nB, nW = len(primaries), len(backups), len(scenarios)

    # Supplier attribute arrays
    cap_p  = np.array([r[1] for r in prim], dtype=float)
    cost_p = np.array([r[2] for r in prim], dtype=float)
    soc_p  = np.array([r[3] for r in prim], dtype=float)
    env_p  = np.array([r[4] for r in prim], dtype=float)
    cap_b  = np.array([r[1] for r in back], dtype=float)
    cost_b = np.array([r[2] for r in back], dtype=float)
    soc_b  = np.array([r[3] for r in back], dtype=float)
    env_b  = np.array([r[4] for r in back], dtype=float)

    # Disruption matrix: dis[s, w] = 1 if supplier s is disrupted in scenario w
    # Loaded from .lp file (disrupted/2 facts)
    dis = np.zeros((nP, nW), dtype=int)
    for s, w in re.findall(r'disrupted\((\w+),\s*(\w+)\)', lp_txt):
        if s in pi and w in si:
            dis[pi[s], si[w]] = 1

    # Policy parameters from Excel 'Policy Parameters' sheet
    ws_pol = wb['Policy Parameters']
    pol    = {}
    for row in ws_pol.iter_rows(min_row=1, max_row=14, values_only=True):
        if row[0] and row[2] and isinstance(row[2], (int, float)):
            pol[str(row[0])] = row[2]

    demand     = pol.get('Total Demand')
    soc_thresh = pol.get('Social Threshold')
    env_thresh = pol.get('Environmental Threshold')

    # .lp file overrides for thresholds (used when .lp and Excel differ)
    sm = re.search(r'soc_threshold\((\d+)\)', lp_txt)
    em = re.search(r'env_threshold\((\d+)\)', lp_txt)
    if sm: soc_thresh = int(sm.group(1))
    if em: env_thresh = int(em.group(1))

    return dict(
        primaries=primaries, backups=backups, scenarios=scenarios,
        cap_p=cap_p,  cost_p=cost_p, soc_p=soc_p, env_p=env_p,
        cap_b=cap_b,  cost_b=cost_b, soc_b=soc_b, env_b=env_b,
        dis=dis,
        demand=demand, soc_thresh=soc_thresh, env_thresh=env_thresh,
        nP=nP, nB=nB, nW=nW
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — MILP SOLVER
# Implements Equations (1)-(8) from Section 4.1
#
# Variable layout (N = 2*(nP + nB) variables total):
#   [x_0 .. x_{nP-1}]    binary   — primary supplier selected?   (x_s^ω, Eq. 4)
#   [y_0 .. y_{nB-1}]    binary   — backup  supplier selected?   (y_b^ω, Eq. 4)
#   [q_0 .. q_{nP-1}]    ≥ 0      — units allocated from primary (q_s^ω)
#   [qb_0.. qb_{nB-1}]   ≥ 0      — units allocated from backup  (q_b^ω)
# ─────────────────────────────────────────────────────────────────────────────

def solve_one_scenario(data, demand, soc_thresh, env_thresh, scenario_idx):
    """
    Solve the MILP for one (demand, disruption scenario) combination.

    Objective : minimise total procurement cost  [Eq. (1), alpha=1, beta=gamma=0]
                ESG objectives are hard constraints, not in the objective.
                This mirrors the ASP lexicographic structure:
                  Priority 1: ESG >= threshold  (hard constraint)
                  Priority 2: minimise cost     (objective)

    Constraints:
        (1) Demand satisfaction    : sum(q_s) + sum(q_b) >= D
        (2) Primary capacity       : q_s <= c̃_sω · x_s    [Extension i]
        (3) Backup  capacity       : q_b <= k_b  · y_b
        (6) Social  threshold      : sum(Soc_s·x_s) + sum(Soc_b·y_b) >= SC
        (7) Env.    threshold      : sum(Env_s·x_s) + sum(Env_b·y_b) >= EN
    """

    # Apply disruption: effective capacity = 0 for disrupted suppliers
    # Implements Extension (i): c̃_sω = k_s · (1 − δ_sω)
    cap_p = data['cap_p'].copy()
    cap_p[data['dis'][:, scenario_idx] == 1] = 0.0

    cap_b  = data['cap_b']
    cost_p, cost_b = data['cost_p'], data['cost_b']
    soc_p,  soc_b  = data['soc_p'],  data['soc_b']
    env_p,  env_b  = data['env_p'],  data['env_b']
    nP, nB         = len(cap_p),     len(cap_b)

    # Variable index slices
    ix  = slice(0,          nP)           # x_s  (binary, primary selection)
    iy  = slice(nP,         nP + nB)      # y_b  (binary, backup  selection)
    iqp = slice(nP + nB,    2*nP + nB)   # q_s  (continuous, primary allocation)
    iqb = slice(2*nP + nB,  2*(nP + nB)) # q_b  (continuous, backup  allocation)
    N   = 2 * (nP + nB)

    # Objective: minimise procurement cost only
    c       = np.zeros(N)
    c[iqp]  = cost_p    # cost of allocating from primary suppliers
    c[iqb]  = cost_b    # cost of allocating from backup  suppliers

    # Integrality: x_s and y_b are binary (0 or 1)
    integrality         = np.zeros(N)
    integrality[ix]     = 1
    integrality[iy]     = 1

    # Variable bounds
    lb      = np.zeros(N)       # all variables >= 0
    ub      = np.ones(N)        # binary vars: upper bound = 1
    ub[iqp] = cap_p             # q_s <= c̃_sω (0 if disrupted)  [Eq. (2)]
    ub[iqb] = cap_b             # q_b <= k_b                     [Eq. (3)]

    # Build constraint matrix in COO format, convert to CSR for the solver
    rows_idx, cols_idx, vals_list = [], [], []
    rlo, rhi = [], []

    def add_constraint(col_indices, col_values, lo, hi):
        r = len(rlo)
        rows_idx.extend([r] * len(col_indices))
        cols_idx.extend(col_indices)
        vals_list.extend(col_values)
        rlo.append(lo)
        rhi.append(hi)

    # Constraint (1): demand satisfaction — sum of all allocations >= D
    add_constraint(
        list(range(nP + nB, N)),
        [1.0] * (nP + nB),
        float(demand), np.inf
    )

    # Constraint (2): q_s <= cap_p[s] · x_s  for each primary supplier
    # If supplier is disrupted (cap_p[s] = 0), force x_s = 0 directly
    for s in range(nP):
        if cap_p[s] > 0:
            add_constraint(
                [nP + nB + s, s],
                [1.0, -cap_p[s]],
                -np.inf, 0.0
            )
        else:
            # Disrupted supplier: force selection variable to 0
            add_constraint([s], [1.0], 0.0, 0.0)

    # Constraint (3): q_b <= cap_b[b] · y_b  for each backup supplier
    for b in range(nB):
        add_constraint(
            [2*nP + nB + b, nP + b],
            [1.0, -cap_b[b]],
            -np.inf, 0.0
        )

    # Constraint (6): aggregate social score >= SC threshold
    add_constraint(
        list(range(nP)) + list(range(nP, nP + nB)),
        list(soc_p)     + list(soc_b),
        float(soc_thresh), np.inf
    )

    # Constraint (7): aggregate environmental score >= EN threshold
    add_constraint(
        list(range(nP)) + list(range(nP, nP + nB)),
        list(env_p)     + list(env_b),
        float(env_thresh), np.inf
    )

    # Assemble sparse constraint matrix and call HiGHS solver
    A = csr_matrix(
        (vals_list, (rows_idx, cols_idx)),
        shape=(len(rlo), N)
    )

    t_start = time.perf_counter()
    result  = milp(
        c,
        constraints = LinearConstraint(A, rlo, rhi),
        integrality = integrality,
        bounds      = Bounds(lb, ub),
        options     = {'disp': False, 'time_limit': 600}
    )
    elapsed = time.perf_counter() - t_start

    # Extract solution metrics
    if result.success:
        x  = result.x
        qp = x[iqp];  qb = x[iqb]
        xs = x[ix];   yb = x[iy]

        total_supply = float(np.sum(qp) + np.sum(qb))

        return dict(
            feasible      = True,
            n_primary_sel = int(np.sum(xs > 0.5)),
            n_backups     = int(np.sum(yb > 0.5)),
            total_cost    = round(float(np.dot(cost_p, qp) + np.dot(cost_b, qb)), 2),
            total_soc     = round(float(np.dot(soc_p,  xs) + np.dot(soc_b,  yb)), 2),
            total_env     = round(float(np.dot(env_p,  xs) + np.dot(env_b,  yb)), 2),
            supply        = round(total_supply, 1),
            fill_rate     = round(min(100.0, 100.0 * total_supply / demand), 2),
            time_s        = round(elapsed, 4)
        )

    # Infeasible or time-limit exceeded
    return dict(
        feasible=False, n_primary_sel=None, n_backups=None,
        total_cost=None, total_soc=None, total_env=None,
        supply=None, fill_rate=None, time_s=round(elapsed, 4)
    )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — EXPERIMENT RUNNER
# ─────────────────────────────────────────────────────────────────────────────

# Load all three datasets
# Each scale: Excel file (supplier attributes) + .lp file (disruption facts)
print("Loading datasets from Figshare files...")
d100  = parse_excel_lp('dataset_100.xlsx',  'instance_100.lp')
d500  = parse_excel_lp('dataset_500.xlsx',  'instance_500.lp')
d1000 = parse_excel_lp('dataset_1000.xlsx', 'instance_1000.lp')

print(f"100-supplier : nP={d100['nP']:>4}, nB={d100['nB']:>3}, "
      f"soc_thresh={d100['soc_thresh']}, env_thresh={d100['env_thresh']}")
print(f"500-supplier : nP={d500['nP']:>4}, nB={d500['nB']:>3}, "
      f"soc_thresh={d500['soc_thresh']}, env_thresh={d500['env_thresh']}")
print(f"1000-supplier: nP={d1000['nP']:>4}, nB={d1000['nB']:>3}, "
      f"soc_thresh={d1000['soc_thresh']}, env_thresh={d1000['env_thresh']}")

# Demand points from paper (Table 4) — used for all MILP runs
PAPER_DEMANDS = {
    100:  [8000,  10000,  12000],
    500:  [40000, 45000,  50000],
    1000: [90000, 95000, 100000],
}

# Published ASP backup counts from Tables 4-5 of the paper
# Used only for comparison in the output — not used in computation
ASP_BACKUPS = {
    (100,   8000, 'w1'):  0,  (100,   8000, 'w2'):  0,  (100,   8000, 'w3'):  0,
    (100,  10000, 'w1'):  8,  (100,  10000, 'w2'):  7,  (100,  10000, 'w3'):  8,
    (100,  12000, 'w1'): 17,  (100,  12000, 'w2'): 16,  (100,  12000, 'w3'): 19,
    (500,  40000, 'w1'):  0,  (500,  40000, 'w2'):  0,  (500,  40000, 'w3'):  0,
    (500,  45000, 'w1'):  3,  (500,  45000, 'w2'):  0,  (500,  45000, 'w3'):  5,
    (500,  50000, 'w1'): 28,  (500,  50000, 'w2'): 25,  (500,  50000, 'w3'): 31,
    (1000, 90000, 'w1'):  0,  (1000, 90000, 'w2'):  0,  (1000, 90000, 'w3'):  0,
    (1000, 95000, 'w1'): 21,  (1000, 95000, 'w2'): 24,  (1000, 95000, 'w3'): 20,
    (1000,100000, 'w1'): 47,  (1000,100000, 'w2'): 52,  (1000,100000, 'w3'): 46,
}

# Run MILP on all 27 instances
all_results = []

print()
print(f"{'Scale':>6} {'Demand':>7} {'Sc':>4} | "
      f"{'Feas':>5} {'Cost':>12} {'Soc':>8} {'Env':>8} "
      f"{'M_Bk':>5} {'Fill%':>6} {'M_t(s)':>8} | "
      f"{'A_Bk':>5} Notes")
print("-" * 100)

for data, scale in [(d100, 100), (d500, 500), (d1000, 1000)]:

    scenarios = data['scenarios']                    # ['w1', 'w2', 'w3']
    si        = {w: i for i, w in enumerate(scenarios)}

    for demand in PAPER_DEMANDS[scale]:
        for scen in scenarios:

            result = solve_one_scenario(
                data,
                demand       = demand,
                soc_thresh   = data['soc_thresh'],
                env_thresh   = data['env_thresh'],
                scenario_idx = si[scen]
            )

            asp_bk = ASP_BACKUPS.get((scale, demand, scen), None)

            # Build comparison note
            if result['feasible'] and asp_bk is not None:
                note = f"MILP={result['n_backups']}  ASP={asp_bk}"
            else:
                note = "—"

            # Append full result row
            all_results.append(dict(
                scale=scale, demand=demand, scenario=scen,
                asp_backups=asp_bk,
                **result
            ))

            # Print row
            cost_s = f"{result['total_cost']:>12.1f}" if result['feasible'] else "    INFEASIBLE"
            soc_s  = f"{result['total_soc']:>8.1f}"   if result['feasible'] else "       —"
            env_s  = f"{result['total_env']:>8.1f}"   if result['feasible'] else "       —"
            bk_s   = f"{result['n_backups']:>5}"      if result['feasible'] else "    —"
            fill_s = f"{result['fill_rate']:>6}"       if result['feasible'] else "     —"

            print(f"{scale:>6} {demand:>7} {scen:>4} | "
                  f"{'Y' if result['feasible'] else 'N':>5} "
                  f"{cost_s} {soc_s} {env_s} "
                  f"{bk_s} {fill_s} {result['time_s']:>8.4f} | "
                  f"{str(asp_bk):>5} {note}")


# Save results to JSON
with open('milp_results.json', 'w') as f:
    json.dump(all_results, f, indent=2)

print()
print("Done. Results saved to milp_results.json")
print(f"Total instances run: {len(all_results)}")
print(f"Feasible: {sum(1 for r in all_results if r['feasible'])}")
print(f"Infeasible: {sum(1 for r in all_results if not r['feasible'])}")
